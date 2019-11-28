#!/usr/bin/env python3
# -*-encoding: utf-8-*-
# author: Valentyn Kofanov


#!/usr/bin/env python3
# -*-encoding: utf-8-*-
# author: Valentyn Kofanov


from kivy.lang import Builder
from kivy.uix.screenmanager import Screen
import openpyxl
from src.Positions import PositionsDB
from data.const import PATH_DB


Builder.load_file('./kv/saveposition.kv')


class SavePosition(Screen):

    def save(self):
        path = f'{str(self.save_position_file.path)}/{self.name_input.text}.xlsx'

        pdb = PositionsDB(PATH_DB)
        pdb.create()
        res = pdb.get_all()
        xl = openpyxl.Workbook()
        xl.create_sheet(title="positions")
        xl.active.cell(1, 1).value = "position_id"
        xl.active.cell(1, 2).value = "position_name"
        xl.active.cell(1, 3).value = "salary_ph"
        xl.active.cell(1, 4).value = "number_hours"

        i = 2
        for row in res:
            xl.active.cell(i, 1).value = row[0]
            xl.active.cell(i, 2).value = row[1]
            xl.active.cell(i, 3).value = row[2]
            xl.active.cell(i, 4).value = row[3]
            i += 1
        xl.save(path)


