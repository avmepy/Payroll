#!/usr/bin/env python3
# -*-encoding: utf-8-*-
# author: Valentyn Kofanov


from kivy.lang import Builder
from kivy.uix.screenmanager import Screen
import openpyxl
from src.Workers import WorkersDB
from data.const import PATH_DB


Builder.load_file('./kv/saveworker.kv')


class SaveWorker(Screen):

    def save(self):
        path = f'{str(self.save_worker_file.path)}/{self.name_input.text}.xlsx'

        wdb = WorkersDB(PATH_DB)
        wdb.create()
        res = wdb.get_all()
        xl = openpyxl.Workbook()
        xl.create_sheet(title="workers")
        xl.active.cell(1, 1).value = "worker_id"
        xl.active.cell(1, 2).value = "name"
        xl.active.cell(1, 3).value = "position_id"

        i = 2
        for row in res:
            xl.active.cell(i, 1).value = row[0]
            xl.active.cell(i, 2).value = row[1]
            xl.active.cell(i, 3).value = row[2]
            i += 1
        xl.save(path)


