#!/usr/bin/env python3
# -*-encoding: utf-8-*-
# author: Valentyn Kofanov


#!/usr/bin/env python3
# -*-encoding: utf-8-*-
# author: Valentyn Kofanov


from kivy.lang import Builder
from kivy.uix.screenmanager import Screen
import openpyxl
from src.ReportCard import ReportCartDB
from data.const import PATH_DB


Builder.load_file('./kv/savepay.kv')


class SavePay(Screen):

    def save(self):
        path = f'{str(self.save_pay_file.path)}/{self.name_input.text}.xlsx'

        rdb = ReportCartDB(PATH_DB)
        rdb.create()
        res = rdb.get_all()
        xl = openpyxl.Workbook()
        xl.create_sheet(title="reports")
        xl.active.cell(1, 1).value = "id_report"
        xl.active.cell(1, 2).value = "id_worker"
        xl.active.cell(1, 3).value = "month"
        xl.active.cell(1, 4).value = "amount"

        i = 2
        for row in res:
            xl.active.cell(i, 1).value = row[0]
            xl.active.cell(i, 2).value = row[1]
            xl.active.cell(i, 3).value = row[2]
            xl.active.cell(i, 4).value = row[3]
            i += 1
        xl.save(path)


